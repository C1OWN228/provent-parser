import json
import re
import html
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse, parse_qs

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import yt_dlp


# ============================================================
# PROVENT PARSER
# ============================================================

# ---------- НАСТРОЙКИ ПОЛЬЗОВАТЕЛЯ ----------

CATALOG_URL = "https://provent.ru/ventustanovki-alpha/"

# Сколько товаров обрабатывать:
# None = все товары каталога
# 1 = только первый
# 5 = первые пять
MAX_PRODUCTS = 1

# Сколько страниц каталога максимум просматривать.
# None = пока не закончатся новые товары.
MAX_CATALOG_PAGES = None

# Папка со всеми результатами
RESULTS_DIR = Path.cwd() / "Результаты"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/148.0.0.0 Safari/537.36"
    )
}

TIMEOUT = 40

# Сколько раз пробовать один товар при временной ошибке.
# 3 = три попытки подряд, прежде чем временно поместить товар в список ошибок.
MAX_ATTEMPTS = 3

# Пауза между попытками, секунд.
RETRY_DELAY = 3


# ============================================================
# СЛУЖЕБНЫЕ ФУНКЦИИ
# ============================================================

def clean_text(text):
    return " ".join(str(text).split()).strip()


def extract_characteristics(description_panel):
    """Извлекает обычные характеристики и широкие таблицы отдельно."""
    characteristics = {}
    characteristic_tables = []

    if not description_panel:
        return characteristics, characteristic_tables

    for li in description_panel.find_all("li"):
        text = clean_text(li.get_text())
        if ":" not in text:
            continue
        key, value = map(clean_text, text.split(":", 1))
        if key and value and len(key) <= 150 and len(value) <= 500:
            characteristics[key] = value

    for table in description_panel.find_all("table"):
        table_rows = []

        for row in table.find_all("tr"):
            cells = row.find_all(["th", "td"], recursive=False)
            if not cells:
                cells = row.find_all(["th", "td"])

            values = [
                clean_text(cell.get_text(" ", strip=True))
                for cell in cells
            ]
            values = [value for value in values if value]
            if len(values) >= 2:
                table_rows.append(values)

        if not table_rows:
            continue

        # Обычные таблицы «параметр — значение» оставляем в характеристиках.
        if not any(len(row) > 2 for row in table_rows):
            for row in table_rows:
                characteristics[row[0]] = " ".join(row[1:])
            continue

        # Широкие таблицы нельзя склеивать в одну строку: сохраняем все ячейки.
        title_element = table.find_previous(["h2", "h3", "h4", "h5", "h6"])
        title = clean_text(title_element.get_text()) if title_element else ""
        if not title:
            title = f"Таблица характеристик {len(characteristic_tables) + 1}"

        characteristic_tables.append({
            "title": title,
            "rows": table_rows,
        })

    for dl in description_panel.find_all("dl"):
        terms = dl.find_all("dt")
        values = dl.find_all("dd")
        for term, value in zip(terms, values):
            key = clean_text(term.get_text())
            val = clean_text(value.get_text())
            if key and val:
                characteristics[key] = val

    return characteristics, characteristic_tables


def safe_filename(filename):
    return re.sub(r'[<>:"/\\|?*]', "_", str(filename)).strip(" ._")


def get_extension(url, default=""):
    try:
        ext = Path(urlparse(url).path).suffix.lower()
    except Exception:
        return default
    if ext and len(ext) <= 10:
        return ext
    return default


def is_valid_image(url):
    return get_extension(url) in {".jpg", ".jpeg", ".png", ".webp"}


def is_service_image(img):
    alt = clean_text(img.get("alt", "")).lower()
    classes = " ".join(img.get("class", [])).lower()
    info = f"{alt} {classes}"
    forbidden = [
        "logo", "логотип", "icon", "икон", "pin", "marker",
        "location", "карта", "avatar", "social"
    ]
    return any(word in info for word in forbidden)


def get_image_url(img, page_url):
    src = (
        img.get("data-large_image")
        or img.get("data-src")
        or img.get("data-lazy-src")
        or img.get("data-original")
        or img.get("src")
    )
    if not src:
        return None
    src = urljoin(page_url, src)
    return re.sub(r"-\d+x\d+(?=\.[a-zA-Z]+$)", "", src)


def download_file(url, folder, filename):
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        path = folder / filename
        path.write_bytes(r.content)
        return path
    except Exception:
        return None


def print_progress(done, total, article="", status=""):
    if total:
        prefix = f"[{done}/{total}]"
    else:
        prefix = f"[{done}]"
    text = f"{prefix} {article}"
    if status:
        text += f" — {status}"
    print(text)


# ============================================================
# VK VIDEO
# ============================================================

def clean_video_url(url):
    if not url:
        return None
    url = html.unescape(url)
    url = url.replace("\\u0026", "&")
    url = url.replace("&amp;", "&")
    url = url.replace("\\/", "/")
    return url.rstrip("\\")


def find_vk_video_urls(page_html, product_root, page_url):
    text = html.unescape(page_html)
    text = text.replace("\\u0026", "&").replace("\\/", "/")

    found = []

    pattern = re.compile(
        r'https?://(?:vk\.com|vkvideo\.ru)/video_ext\.php'
        r'\?[^"\'<>\s\\]+',
        re.IGNORECASE
    )
    for url in pattern.findall(text):
        url = clean_video_url(url)
        if url and url not in found:
            found.append(url)

    pattern = re.compile(
        r'https?://vkvideo\.ru/video(?:-[\w-]+)?_\d+',
        re.IGNORECASE
    )
    for url in pattern.findall(text):
        url = clean_video_url(url)
        if url and url not in found:
            found.append(url)

    for iframe in product_root.find_all("iframe"):
        src = iframe.get("src")
        if not src:
            continue
        src = clean_video_url(urljoin(page_url, src))
        if src and ("vkvideo" in src.lower() or "vk.com/video" in src.lower()):
            if src not in found:
                found.append(src)

    return found


def convert_to_vkvideo(url):
    url = clean_video_url(url)
    if not url:
        return None

    if "video_ext.php" in url.lower():
        parsed = urlparse(url)
        query = parse_qs(parsed.query)
        oid = query.get("oid", [None])[0]
        video_id = query.get("id", [None])[0]
        if oid and video_id:
            return f"https://vkvideo.ru/video{oid}_{video_id}"

    match = re.search(
        r"(https?://vkvideo\.ru/video(?:-[\w-]+)?_\d+)",
        url,
        re.IGNORECASE
    )
    return match.group(1) if match else None


def download_vk_videos(vk_urls, videos_folder, article):
    downloaded = []

    for number, vk_url in enumerate(vk_urls, 1):
        template = str(videos_folder / f"{article}_видео_{number}.%(ext)s")

        opts = {
            "format": "bestvideo*+bestaudio/best",
            "outtmpl": template,
            "noplaylist": True,
            "merge_output_format": "mp4",
            "quiet": True,
            "no_warnings": True,
            "overwrites": True,
            "http_headers": HEADERS,
        }

        try:
            with yt_dlp.YoutubeDL(opts) as ydl:
                ydl.extract_info(vk_url, download=True)

            candidates = [
                p for p in videos_folder.glob(f"{article}_видео_{number}.*")
                if p.suffix.lower() not in {".part", ".ytdl"}
            ]
            if candidates:
                downloaded.append(candidates[0].name)
        except Exception:
            pass

    return downloaded



def canonical_url(url):
    """Убирает query/fragment, чтобы одна и та же страница не считалась разными товарами."""
    parsed = urlparse(url)
    return parsed._replace(query="", fragment="").geturl().rstrip("/") + "/"


def is_product_link(href, catalog_url):
    """Проверяет, похожа ли ссылка на страницу товара внутри текущего каталога."""
    if not href:
        return False

    full = canonical_url(urljoin(catalog_url, href))
    base = canonical_url(catalog_url)

    if not full.startswith(base):
        return False

    parsed = urlparse(full)
    path = parsed.path.rstrip("/")

    base_path = urlparse(base).path.rstrip("/")
    relative = path[len(base_path):].strip("/") if path.startswith(base_path) else ""

    if not relative:
        return False

    # Не считаем страницами товаров пагинацию и служебные URL.
    parts = [p for p in relative.split("/") if p]
    forbidden = {
        "page", "tag", "category", "product-category",
        "filter", "feed", "cart", "checkout", "my-account"
    }

    if any(part.lower() in forbidden for part in parts):
        return False

    if re.search(r"/page/\d+$", path, re.I):
        return False

    # Для PROVENT товар находится глубже категории.
    return len(parts) >= 1


def extract_catalog_product_urls(catalog_url):
    """
    Обходит страницы каталога и собирает уникальные URL товаров.
    Пагинация определяется по ссылкам /page/N/ и rel=next.
    """
    session = requests.Session()
    session.headers.update(HEADERS)

    base = canonical_url(catalog_url)
    queue = [base]
    visited_pages = set()
    product_urls = []
    seen_products = set()

    while queue:
        page_url = queue.pop(0)
        page_url = canonical_url(page_url)

        if page_url in visited_pages:
            continue

        if MAX_CATALOG_PAGES is not None and len(visited_pages) >= MAX_CATALOG_PAGES:
            break

        visited_pages.add(page_url)

        try:
            response = session.get(page_url, timeout=TIMEOUT)
            response.raise_for_status()
        except Exception as error:
            print(f"  Не удалось открыть страницу каталога: {page_url}")
            print(f"  Причина: {error}")
            continue

        soup = BeautifulSoup(response.text, "html.parser")

        # Сначала ищем ссылки внутри карточек WooCommerce.
        candidates = []

        for selector in [
            "li.product a[href]",
            ".products a[href]",
            ".product a[href]",
            ".product-item a[href]",
            ".product-card a[href]",
        ]:
            candidates.extend(soup.select(selector))

        # Если структура сайта поменяется — дополнительно смотрим все ссылки.
        if not candidates:
            candidates = soup.find_all("a", href=True)

        for link in candidates:
            href = link.get("href")
            if not is_product_link(href, base):
                continue

            product_url = canonical_url(urljoin(page_url, href))

            if product_url == base:
                continue

            if product_url not in seen_products:
                seen_products.add(product_url)
                product_urls.append(product_url)

                # Если нужен только небольшой тестовый объём,
                # дальше собирать каталог нет смысла.
                if MAX_PRODUCTS is not None and len(product_urls) >= MAX_PRODUCTS:
                    return product_urls

        # Ищем следующую страницу каталога.
        next_urls = []

        for link in soup.find_all("a", href=True):
            href = link.get("href", "")
            rel = link.get("rel", [])
            rel_text = " ".join(rel) if isinstance(rel, list) else str(rel)
            label = clean_text(link.get_text()).lower()

            if "next" in rel_text.lower() or label in {
                "следующая", "далее", "next", ">", "»"
            }:
                next_url = canonical_url(urljoin(page_url, href))
                if next_url.startswith(base) and next_url not in visited_pages:
                    next_urls.append(next_url)

            if re.search(r"/page/\d+/?$", urlparse(urljoin(page_url, href)).path, re.I):
                next_url = canonical_url(urljoin(page_url, href))
                if next_url.startswith(base) and next_url not in visited_pages:
                    next_urls.append(next_url)

        for next_url in next_urls:
            if next_url not in queue and next_url not in visited_pages:
                queue.append(next_url)

    return product_urls


def build_product_signature(response_text, soup, product_root, article, url):
    """
    Сигнатура товара — компактный отпечаток того, что важно для парсера.
    Если описание, цена, характеристики, картинки, документы или VK-видео
    изменились — хэш изменится и товар будет перепарсен.
    """
    root_text = clean_text(product_root.get_text(" ", strip=True))

    images = []
    for img in product_root.find_all("img"):
        image_url = get_image_url(img, url)
        if image_url and image_url not in images:
            images.append(image_url)

    files = []
    file_extensions = {
        ".pdf", ".zip", ".rar", ".doc", ".docx", ".xls", ".xlsx",
        ".dwg", ".dxf", ".rvt", ".ifc", ".stp", ".step",
        ".igs", ".iges"
    }
    for link in product_root.find_all("a", href=True):
        href = urljoin(url, link.get("href"))
        if get_extension(href) in file_extensions and href not in files:
            files.append(href)

    videos = find_vk_video_urls(response_text, product_root, url)

    state = {
        "url": canonical_url(url),
        "article": article,
        "text": root_text,
        "images": sorted(images),
        "files": sorted(files),
        "videos": sorted(videos),
    }

    raw = json.dumps(state, ensure_ascii=False, sort_keys=True)
    import hashlib
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def inspect_product(url):
    """
    Быстрая проверка товара без скачивания файлов.
    Нужна, чтобы повторный запуск не перекачивал неизменившиеся товары.
    """
    response = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    h1 = soup.find("h1")

    product_root = None
    if h1:
        current = h1
        for _ in range(12):
            if not current.parent:
                break
            current = current.parent
            text = clean_text(current.get_text(" ", strip=True))
            if "Артикул:" in text and "Описание" in text:
                product_root = current
                break

    if product_root is None:
        product_root = soup

    product_text = clean_text(product_root.get_text(" ", strip=True))
    article_match = re.search(
        r"Артикул\s*:\s*([A-Za-zА-Яа-я0-9._/-]+)",
        product_text
    )
    article = article_match.group(1) if article_match else "unknown"

    signature = build_product_signature(
        response.text, soup, product_root, article, url
    )

    return {
        "article": article,
        "signature": signature,
    }


def product_needs_update(product_folder, article, current_signature):
    """Возвращает True, если товара ещё нет или его содержимое изменилось."""
    json_path = product_folder / f"{article}.json"

    if not json_path.exists():
        return True

    try:
        old_data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return True

    old_signature = old_data.get("_source_signature")

    # Старые результаты, созданные предыдущей версией парсера,
    # один раз перепарсим, чтобы добавить сигнатуру.
    if not old_signature:
        return True

    if old_signature != current_signature:
        return True

    # Если VK-ссылка есть, но файл видео пропал/не скачался,
    # не считаем товар полностью готовым.
    downloaded_videos = old_data.get("downloaded_videos", [])
    if downloaded_videos:
        videos_folder = product_folder / "Видео"
        for filename in downloaded_videos:
            if not (videos_folder / filename).exists():
                return True

    return False


# ============================================================
# ПАРСИНГ ТОВАРА
# ============================================================

def parse_product(url, source_signature=None):
    response = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    h1 = soup.find("h1")
    name = clean_text(h1.get_text()) if h1 else "Без названия"

    product_root = None
    if h1:
        current = h1
        for _ in range(12):
            if not current.parent:
                break
            current = current.parent
            text = clean_text(current.get_text(" ", strip=True))
            if "Артикул:" in text and "Описание" in text:
                product_root = current
                break

    if product_root is None:
        product_root = soup

    product_text = clean_text(product_root.get_text(" ", strip=True))
    article_match = re.search(
        r"Артикул\s*:\s*([A-Za-zА-Яа-я0-9._/-]+)",
        product_text
    )
    article = article_match.group(1) if article_match else "unknown"

    # ---------- ПАПКИ ----------
    product_folder = RESULTS_DIR / safe_filename(article)
    gallery_folder = product_folder / "Галерея"
    description_folder = product_folder / "Описание"
    characteristics_folder = product_folder / "Характеристики"
    downloads_folder = product_folder / "Загрузки"

    for folder in [
        gallery_folder,
        description_folder,
        characteristics_folder,
        downloads_folder,
    ]:
        folder.mkdir(parents=True, exist_ok=True)

    # ---------- ЦЕНА ----------
    price = None
    for selector in [
        ".price",
        ".product-price",
        ".woocommerce-Price-amount",
        "[class*='price']",
    ]:
        for element in product_root.select(selector):
            match = re.search(r"(\d[\d\s]*)\s*₽", clean_text(element.get_text()))
            if match and match.group(1).replace(" ", "") != "0":
                price = match.group(1).strip() + " ₽"
                break
        if price:
            break

    if price is None:
        for element in product_root.find_all(string=re.compile(r"₽")):
            match = re.search(r"(\d[\d\s]*)\s*₽", clean_text(element))
            if match and match.group(1).replace(" ", "") != "0":
                price = match.group(1).strip() + " ₽"
                break

    # ---------- ОПИСАНИЕ ----------
    description_panel = None
    for selector in [
        ".woocommerce-Tabs-panel--description",
        "#tab-description",
        "[id*='description']",
    ]:
        candidate = soup.select_one(selector)
        if candidate:
            description_panel = candidate
            break

    if description_panel is None:
        for link in soup.find_all("a"):
            if clean_text(link.get_text()).lower() == "описание":
                parent = link.parent
                for _ in range(8):
                    if not parent:
                        break
                    text = clean_text(parent.get_text(" ", strip=True))
                    if "Загрузки" in text and len(text) > 100:
                        description_panel = parent
                        break
                    parent = parent.parent
                if description_panel:
                    break

    description_parts = []
    if description_panel:
        for p in description_panel.find_all("p"):
            text = clean_text(p.get_text())
            if text:
                description_parts.append(text)

        if not description_parts:
            for div in description_panel.find_all("div"):
                if div.find(["p", "ul", "ol", "table"]):
                    continue
                text = clean_text(div.get_text())
                if len(text) >= 10:
                    description_parts.append(text)

    description = "\n\n".join(dict.fromkeys(description_parts))

    # ---------- ХАРАКТЕРИСТИКИ ----------
    characteristics, characteristic_tables = extract_characteristics(
        description_panel
    )

    # ---------- ИЗОБРАЖЕНИЯ ----------
    gallery_images = []
    description_images = []
    characteristics_images = []

    gallery = soup.select_one(".woocommerce-product-gallery")
    if gallery:
        for button in gallery.select(".product-detail-gallery__thumb"):
            style = button.get("style", "")
            match = re.search(r"url\(['\"]?([^'\")]+)['\"]?\)", style)
            if not match:
                continue
            image_url = re.sub(
                r"-\d+x\d+(?=\.[a-zA-Z]+$)", "",
                urljoin(url, match.group(1))
            )
            if is_valid_image(image_url) and image_url not in gallery_images:
                gallery_images.append(image_url)

        main_image = gallery.select_one(".product-detail-gallery__main img")
        if main_image:
            image_url = get_image_url(main_image, url)
            if image_url and is_valid_image(image_url) and image_url not in gallery_images:
                gallery_images.insert(0, image_url)

    characteristic_heading = None
    if description_panel:
        for heading in description_panel.find_all(["h2", "h3", "h4", "h5", "h6"]):
            if "характеристик" in clean_text(heading.get_text()).lower():
                characteristic_heading = heading
                break

        for img in description_panel.find_all("img"):
            if is_service_image(img):
                continue
            image_url = get_image_url(img, url)
            if not image_url or not is_valid_image(image_url):
                continue

            is_characteristic = False
            if characteristic_heading:
                for element in characteristic_heading.find_all_next():
                    if element is img:
                        is_characteristic = True
                        break
                    if (
                        element.name in {"h2", "h3", "h4", "h5", "h6"}
                        and element is not characteristic_heading
                    ):
                        break

            target = characteristics_images if is_characteristic else description_images
            if image_url not in target:
                target.append(image_url)

    # ---------- СКАЧИВАНИЕ КАРТИНОК ----------
    downloaded_gallery = []
    for i, image_url in enumerate(gallery_images, 1):
        ext = get_extension(image_url, ".jpg")
        filename = f"{article}_фото_{i}{ext}"
        if download_file(image_url, gallery_folder, filename):
            downloaded_gallery.append(filename)

    downloaded_description = []
    for i, image_url in enumerate(description_images, 1):
        ext = get_extension(image_url, ".jpg")
        filename = f"{article}_описание_{i}{ext}"
        if download_file(image_url, description_folder, filename):
            downloaded_description.append(filename)

    downloaded_characteristics = []
    for i, image_url in enumerate(characteristics_images, 1):
        ext = get_extension(image_url, ".jpg")
        filename = f"{article}_характеристики_{i}{ext}"
        if download_file(image_url, characteristics_folder, filename):
            downloaded_characteristics.append(filename)

    # ---------- ЗАГРУЗКИ ----------
    downloads = []
    file_extensions = {
        ".pdf", ".zip", ".rar", ".doc", ".docx", ".xls", ".xlsx",
        ".dwg", ".dxf", ".rvt", ".ifc", ".stp", ".step",
        ".igs", ".iges"
    }

    downloads_panel = None
    for selector in [
        ".woocommerce-Tabs-panel--downloads",
        "#tab-downloads",
        "[id*='downloads']",
    ]:
        candidate = soup.select_one(selector)
        if candidate:
            downloads_panel = candidate
            break

    if downloads_panel is None:
        for link in soup.find_all("a"):
            if clean_text(link.get_text()).lower() == "загрузки":
                parent = link.parent
                for _ in range(6):
                    if not parent:
                        break
                    if any(
                        get_extension(a.get("href", "")) in file_extensions
                        for a in parent.find_all("a", href=True)
                    ):
                        downloads_panel = parent
                        break
                    parent = parent.parent
                if downloads_panel:
                    break

    if downloads_panel:
        for link in downloads_panel.find_all("a", href=True):
            file_url = urljoin(url, link.get("href"))
            if get_extension(file_url) in file_extensions and file_url not in downloads:
                downloads.append(file_url)

    downloaded_files = []
    for i, file_url in enumerate(downloads, 1):
        ext = get_extension(file_url, ".bin")
        link_text = ""
        if downloads_panel:
            for link in downloads_panel.find_all("a", href=True):
                if urljoin(url, link.get("href")) == file_url:
                    link_text = clean_text(link.get_text())
                    break

        filename = safe_filename(link_text) if link_text else f"download_{i}"
        if not Path(filename).suffix:
            filename += ext
        filename = f"{article}_{filename}"

        if download_file(file_url, downloads_folder, filename):
            downloaded_files.append(filename)

    # ---------- ВИДЕО ----------
    raw_video_urls = find_vk_video_urls(response.text, product_root, url)
    vkvideo_urls = []
    for raw in raw_video_urls:
        normal = convert_to_vkvideo(raw)
        if normal and normal not in vkvideo_urls:
            vkvideo_urls.append(normal)

    downloaded_videos = []
    videos_folder = None

    if vkvideo_urls:
        videos_folder = product_folder / "Видео"
        videos_folder.mkdir(parents=True, exist_ok=True)
        downloaded_videos = download_vk_videos(
            vkvideo_urls, videos_folder, article
        )

    # ---------- JSON ----------
    product_data = {
        "_source_signature": source_signature,
        "name": name,
        "article": article,
        "price": price,
        "description": description,
        "characteristics": characteristics,
        "characteristic_tables": characteristic_tables,
        "source_url": url,
        "gallery": downloaded_gallery,
        "description_images": downloaded_description,
        "characteristics_images": downloaded_characteristics,
        "downloads": downloaded_files,
        "vk_video_urls": vkvideo_urls,
        "downloaded_videos": downloaded_videos,
    }

    json_path = product_folder / f"{article}.json"
    json_path.write_text(
        json.dumps(product_data, ensure_ascii=False, indent=4),
        encoding="utf-8"
    )

    # ---------- EXCEL ----------
    excel_path = product_folder / f"{article}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Товар"

    ws.append(["Поле", "Значение"])
    ws.append(["Название", name])
    ws.append(["Артикул", article])
    ws.append(["Цена", price])
    ws.append(["Описание", description])
    ws.append(["Ссылка на товар", url])
    ws.append([])
    ws.append(["ХАРАКТЕРИСТИКИ", ""])

    if characteristics:
        for key, value in characteristics.items():
            ws.append([key, value])
    else:
        ws.append(["Характеристики", "Не найдены в текстовом виде"])

    if characteristic_tables:
        widest_table = 1

        for table_number, characteristic_table in enumerate(
            characteristic_tables, 1
        ):
            ws.append([])
            ws.append([characteristic_table["title"]])

            for table_row in characteristic_table["rows"]:
                ws.append(table_row)
                widest_table = max(widest_table, len(table_row))

        # Узкие колонки нужны, чтобы значения таблицы были видны рядом.
        ws.column_dimensions["B"].width = 18
        for column_number in range(3, widest_table + 1):
            ws.column_dimensions[get_column_letter(column_number)].width = 14

    ws.append([])
    ws.append(["ФАЙЛЫ", ""])

    for filename in downloaded_gallery:
        ws.append(["Фото", f"Галерея/{filename}"])
    for filename in downloaded_description:
        ws.append(["Изображение описания", f"Описание/{filename}"])
    for filename in downloaded_characteristics:
        ws.append(["Изображение характеристик", f"Характеристики/{filename}"])
    for filename in downloaded_files:
        ws.append(["Загрузка", f"Загрузки/{filename}"])
    for video_url in vkvideo_urls:
        ws.append(["VK Video", video_url])
    for filename in downloaded_videos:
        ws.append(["Скачанное видео", f"Видео/{filename}"])

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    wb.save(excel_path)

    return {
        "article": article,
        "name": name,
        "price": price,
        "folder": product_folder,
        "gallery": len(downloaded_gallery),
        "description_images": len(downloaded_description),
        "characteristics_images": len(downloaded_characteristics),
        "characteristics": len(characteristics),
        "downloads": len(downloaded_files),
        "videos_found": len(vkvideo_urls),
        "videos_downloaded": len(downloaded_videos),
    }


# ============================================================
# ЗАПУСК
# ============================================================

def try_parse_product(url, attempt_label=""):
    """
    Пытается обработать один товар.
    При временной ошибке повторяет запрос несколько раз.
    """
    last_error = None

    for attempt in range(1, MAX_ATTEMPTS + 1):
        try:
            return parse_product(url)
        except Exception as error:
            last_error = error

            if attempt < MAX_ATTEMPTS:
                print(
                    f"    Не удалось обработать товар "
                    f"(попытка {attempt}/{MAX_ATTEMPTS}). "
                    f"Повтор через {RETRY_DELAY} сек..."
                )
                time.sleep(RETRY_DELAY)

    raise last_error


def main():
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 64)
    print("PROVENT PARSER")
    print("=" * 64)
    print(f"Каталог: {CATALOG_URL}")
    print()

    # ========================================================
    # ПОИСК ТОВАРОВ В КАТАЛОГЕ
    # ========================================================

    print("Ищем товары в каталоге...")
    urls = extract_catalog_product_urls(CATALOG_URL)

    found_total = len(urls)

    if not urls:
        print("Товары в каталоге не найдены.")
        return

    print(f"Найдено уникальных товаров: {found_total}")

    if MAX_PRODUCTS is not None:
        urls = urls[:MAX_PRODUCTS]
        print(f"К обработке по лимиту: {len(urls)}")
    else:
        print("К обработке: все найденные товары")

    total = len(urls)
    print()

    results = []
    failed = []
    skipped = 0
    updated = 0
    new_products = 0

    # ========================================================
    # ОСНОВНОЙ ПРОХОД
    # ========================================================

    for index, url in enumerate(urls, 1):
        print(f"[{index}/{total}] Проверяем товар...")

        try:
            # Сначала лёгкая проверка: изменился ли товар.
            state = inspect_product(url)
            article = state["article"]

            product_folder = RESULTS_DIR / safe_filename(article)

            if not product_needs_update(
                product_folder,
                article,
                state["signature"]
            ):
                skipped += 1
                print(
                    f"[{index}/{total}] — {article} — "
                    f"без изменений, пропущен"
                )
                print()
                continue

            # Определяем, новый это товар или обновление.
            existing_json = product_folder / f"{article}.json"
            is_update = existing_json.exists()

            result = try_parse_product(url)
            # try_parse_product вызывает parse_product без сигнатуры,
            # поэтому после полного парсинга дописываем актуальную сигнатуру.
            json_path = product_folder / f"{article}.json"

            try:
                data = json.loads(json_path.read_text(encoding="utf-8"))
                data["_source_signature"] = state["signature"]
                json_path.write_text(
                    json.dumps(data, ensure_ascii=False, indent=4),
                    encoding="utf-8"
                )
            except Exception:
                pass

            results.append(result)

            if is_update:
                updated += 1
                status = "обновлён"
            else:
                new_products += 1
                status = "новый товар"

            print(
                f"[{index}/{total}] ✓ {result['article']} — {status}"
            )
            print(
                f"    Фото: {result['gallery']} | "
                f"Характеристики: {result['characteristics']} | "
                f"Загрузки: {result['downloads']} | "
                f"Видео: {result['videos_downloaded']}/{result['videos_found']}"
            )
            print()

        except Exception as error:
            failed.append({
                "url": url,
                "index": index,
                "error": error,
            })

            print(
                f"[{index}/{total}] ✗ Ошибка — "
                f"будет повторная попытка"
            )
            print()

    # ========================================================
    # ПОВТОР ОШИБОК
    # ========================================================

    if failed:
        print("=" * 64)
        print(f"ПОВТОРНАЯ ПОПЫТКА — {len(failed)} товаров")
        print("=" * 64)
        print()

        still_failed = []

        for item in failed:
            index = item["index"]
            url = item["url"]

            try:
                state = inspect_product(url)
                article = state["article"]
                product_folder = RESULTS_DIR / safe_filename(article)

                result = try_parse_product(url)

                json_path = product_folder / f"{article}.json"
                data = json.loads(json_path.read_text(encoding="utf-8"))
                data["_source_signature"] = state["signature"]
                json_path.write_text(
                    json.dumps(data, ensure_ascii=False, indent=4),
                    encoding="utf-8"
                )

                results.append(result)

                print(
                    f"[{index}/{total}] ✓ {article} — "
                    f"обработан после повтора"
                )
                print()

            except Exception as error:
                still_failed.append({
                    "url": url,
                    "index": index,
                    "error": error,
                })

                print(
                    f"[{index}/{total}] ✗ Не удалось обработать "
                    f"после всех попыток"
                )
                print()

        failed = still_failed

    # ========================================================
    # ИТОГ
    # ========================================================

    print("=" * 64)
    print("ПАРСИНГ ЗАВЕРШЁН")
    print("=" * 64)
    print(f"Найдено в каталоге: {found_total}")
    print(f"Проверено: {total}")
    print(f"Новых товаров: {new_products}")
    print(f"Обновлено: {updated}")
    print(f"Без изменений: {skipped}")
    print(f"Ошибок: {len(failed)}")
    print(f"Результаты: {RESULTS_DIR}")

    if failed:
        print()
        print("Не удалось обработать:")
        for item in failed:
            print(f"  [{item['index']}] {item['url']}")

    print()


if __name__ == "__main__":
    main()
